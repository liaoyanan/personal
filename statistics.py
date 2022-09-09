# -*- coding:utf-8 -*-

import os
import argparse
import datetime
import operator
import openpyxl
from datetime import datetime
from calendar import monthrange
from redminelib import Redmine
from prettytable import PrettyTable
from difflib import SequenceMatcher as SM
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side, colors, Alignment

# 命令行参数
parser = argparse.ArgumentParser(description="Redmine DI Statistics")
parser.add_argument("-a", "--author", type=str, nargs="+", help="Bug reporter, eg. '吴晓飞'")
parser.add_argument("-s", "--sdate", type=str, help="Start date, eg. '2022-04-01'")
parser.add_argument("-d", "--edate", type=str, help="End date, eg. '2022-04-30'")
args = parser.parse_args()


def str_len(strs):
    counter = 0
    for s in strs:
        if '\u4e00' <= s <= '\u9fff':
            counter += 2
        else:
            counter += 1
    return counter+2

class RedmineDI(object):

    def __init__(self, version=None, url=None, key=None):
        """初始化 Redmine 请求地址、身份验证密钥、版本、操作对象等"""
        self.version = version if version else "4.2.3"
        self.url = url if url else 'http://172.16.3.100'
        self.key = key if key else '29b67ad82e398651dd11d3f1a2906d4be4486e46'
        self.redmine = Redmine(url=self.url, key=self.key, version=self.version, raise_attr_exception=False)
        self.adis, self.pdis, self.indivs = {}, {}, {}
        self.di = {
            "P1-P5": 0,     # P1-P5 问题单总数
            "P4-P5": 0,     # P4-P5 问题单总数
            "P1": 0,
            "P2": 0,
            "P3": 0,
            "P4": 0,
            "P5": 0,
            "Invalid": 0,           # 无效问题单总数 (Not Bug & Dubplicate)
            "NoTestcase": 0,        # 随机问题单总数
            "NotReproduced": 0,     # 不复现问题单总数
            "total": 0              # 所有问题单
        }
        # 问题单级别
        self.severity = {"P1": 0.1, "P2": 0.5, "P3": 2, "P4": 5, "P5": 10}
        # 问题单状态
        self.status = {
            "New": 3,       # 新建问题单
            "Open": 4,      # 指派问题单
            "Checked": 8,   # 已处理
            "Fixed": 5,     # 已解决
            "Closed": 2,    # 已关闭
            "Reopen": 1     # 重启问题单
        }
        # 默认项目和部门成员
        self.default_projects = ("CSOS", "云安全", "工业卫士", "态势感知", "监管平台", "神探", "售后支持")
        self.default_team = (
            "廖延安", "韩英杰", "胡建东", "韩非", "李冬生", "范京京", "张斌", "吴学齐", "田杏芝", "程伟磊", "吴晓飞", "周晨",
            "季弘毅", "王九洲", "双鹏鹏", "罗玉", "向菁", "姜慧玲", "刘宇鹏", "董慧敏", "苏珂", "张晨华", "马克", "徐兴基",
            "何东升", "王磊", "张晓光", "刘腾", "马林", "孙伟龙", "张魁", "王谦", '万梦科', '俞忻旺', '孟航程', '张杨', '张毅锋',
            '米嘉豪', '龚家兴',
        )
        # 默认日期
        self.now = datetime.now()
        self.sdate = datetime(self.now.year, self.now.month, 1).strftime("%Y-%m-%d")
        self.edate = datetime(self.now.year, self.now.month, monthrange(self.now.year, self.now.month)[1]).strftime("%Y-%m-%d")

        # 表格数据
        self.team_field_names = [
            "姓 名", "问题单总DI", "问题单总数", "P4-P5总数", "P4-P5比例",
            "无效问题个数", "无效问题比例", "随机问题个数", "随机问题比例", "不复现问题个数", "不复现问题比例"
        ]
        self.product_field_names = self.team_field_names.copy()
        self.product_field_names[0] = "产品线名称"
        self.person, self.project = PrettyTable(), PrettyTable()
        self.f = openpyxl.Workbook()
        self.person.field_names, self.project.field_names = self.team_field_names, self.product_field_names
        self.sheet1, self.sheet2 = self.f.create_sheet(title="产品线统计", index=0),\
                                   self.f.create_sheet(title="测试人员统计", index=1)
        self.sheet1.append(self.product_field_names)
        self.sheet2.append(self.team_field_names)

    def di_statistics(self):
        """统计组内任意成员在任意时间段内的 DI 总值"""

        # 1.确认日期 (默认本月第一天和最后一天)
        if args.sdate and args.edate:
            sdate, edate = args.sdate, args.edate
        else:
            sdate, edate = self.sdate, self.edate
        print("\n统计日期: %s - %s。\n" % (sdate, edate))

        # 2.确认人员 (默认测试部门所有成员,排除非测试部门人员)
        if args.author:
            authors = set(args.author)
            confirm = list(map(lambda member: member in self.default_team, authors))
            members = args.author if all(confirm) else list(filter(lambda m: m in self.default_team, authors))
        else:
            members = self.default_team
        _members = ", ".join(sorted(members, reverse=True))

        # 3.根据查询日期和人员统计数据
        projects = self.redmine.project.all()
        for project in projects:
            if project.name in self.default_projects:
                issues = self.redmine.issue.filter(project_id=project.id, status_id='*', created_on=f"><%s|%s" % (sdate, edate))
                # 统计当前项目的问题单总数
                self.di["total"] = 0
                self.di["total"] += len(issues)

                # 统计当前项目问题单详细数据
                for issue in issues:
                    # 如果提交人名称是管理员或者不在默认名单列表以及跟踪栏目是工时统计时, 则退出本次循环!
                    author = issue.author.name.replace(" ", "").strip()
                    try:
                        name = list(filter(lambda n: SM(None, author, n).quick_ratio() > 0.9, self.default_team))[0]
                    except IndexError:
                        name = author
                    if name not in members or name == "RedmineAdmin" or issue.tracker.name == "工时统计":
                        continue

                    # 检查初始化问题单提交人统计信息[问题单总数、P1-P5总DI值、P4-P5问题单总数、无效问题单总数、不复现或非必现问题单总数]
                    if self.indivs.get(name) is None:
                        self.indivs[name] = {}
                        self.indivs[name]["DI"], self.indivs[name]["Total"], self.indivs[name]["P4-P5"], \
                        self.indivs[name]["P4-P5-DI"], self.indivs[name]["Invalid"],\
                        self.indivs[name]["NoTestcase"], self.indivs[name]["NotReproduced"] = 0, 0, 0, 0, 0, 0, 0
                    self.indivs[name]["Total"] += 1

                    try:
                        # 统计无效问题单数量
                        invaild = issue.custom_fields[2]["value"].lower()
                        if invaild.find("not") >= 0 or invaild.endswith("cate"):
                            self.di["Invalid"] += 1
                            self.indivs[name]["Invalid"] += 1
                            continue

                        # 统计随机问题单数量
                        testcase = issue.custom_fields[0]["value"].lower()
                        if not testcase or testcase.find("无") >= 0 or testcase.find("wu") >= 0 or testcase.find("补充") >= 0:
                            self.di["NoTestcase"] += 1
                            self.indivs[name]["NoTestcase"] += 1

                        # 统计不复现问题单总数 (注: 根据问题单必现情况和解决状态 Can't reproduce 决定)
                        not_reproduced = issue.custom_fields[3]["value"]
                        if not not_reproduced.find("必现") >= 0 or invaild.find("reproduce") >= 0:
                            self.di["NotReproduced"] += 1
                            self.indivs[name]["NotReproduced"] += 1

                        # 统计各级问题单数量 & 统计P1-P5 DI 数值 & 问题单总数 & 统计 P4-P5 DI 数值
                        level = issue.custom_fields[1]["value"][:2]
                        _di = self.severity[level] + dict(self.adis).get(name, 0)
                        self.adis[name], self.indivs[name]["DI"] = _di, _di
                        self.di["P1-P5"] += 1
                        self.di[level] += 1
                        if level.startswith(("P4", "P5")):
                            raw_pdi = dict(self.pdis).get(name, 0)
                            self.pdis[name] = self.severity[level] + raw_pdi
                            self.di["P4-P5"] += 1
                            self.indivs[name]["P4-P5"] += 1

                    except Exception:
                        pass

                # DI数值降序排序
                self.adis = dict(sorted(self.adis.items(), key=operator.itemgetter(1), reverse=True))
                self.pdis = dict(sorted(self.pdis.items(), key=operator.itemgetter(1), reverse=True))

                # 规范 DI 数值精度, 小数点后两位
                for name, di in self.adis.items():
                    self.adis[name] = round(self.adis[name], 1)
                for name, di in self.pdis.items():
                    self.pdis[name] = round(self.pdis[name], 1)

                # 统计数值
                try:
                    # P1-P5
                    P1_P5 = sum(list(self.adis.values()))

                    # P4-P5
                    P4_P5 = sum(list(self.pdis.values()))
                    P4_P5_SCE = round(self.di["P4-P5"] / self.di["P1-P5"] * 100, 2)

                    # P1, P2, P3, P4, P5
                    for p in self.severity:
                        di = self.di[p]
                        di_scale1, di_scale2 = round(di / self.di["P1-P5"] * 100, 2), round((di * self.severity[p]) / P1_P5 * 100, 2)

                    # 无效/随机/不复现问题单
                    invalid_pro = round(self.di["Invalid"] / self.di["total"] * 100, 2)
                    notestcase_pro = round(self.di["NoTestcase"] / self.di["total"] * 100, 2)
                    notre_pro = round(self.di["NotReproduced"] / self.di["total"] * 100, 2)

                    # 项目数据统计结果
                    ndata = [
                        project.name, round(P1_P5, 1), self.di["P1-P5"], self.di["P4-P5"], P4_P5_SCE,
                        self.di["Invalid"], invalid_pro, self.di["NoTestcase"],notestcase_pro, self.di["NotReproduced"], notre_pro
                    ]
                    self.sheet1.append(ndata)
                    self.project.add_row(ndata)
                except Exception:
                    pass

                # 各产品各模块问题单比例
                self.module(project, sdate, edate)

            # 清除历史数据
            self.adis.clear(), self.pdis.clear()
            for key, _ in self.di.items():
                self.di[key] = 0

        # 打印组内成员详细数据
        self.individual()

    def module(self, project, sdate, edate):
        """
        统计任意时间段内各个项目各个模块问题单比例
        :param project: 项目
        :param sdate: 起始日期
        :param edate: 截止日期
        :return: None
        """
        self.moudle = []
        if project.name in self.default_projects:
            minfo = {}
            issues = self.redmine.issue.filter(project_id=project.id, status_id='*', created_on=f"><%s|%s" % (sdate, edate))
            minfo[project.name] = len(issues)
            for issue in issues:
                sub_count = minfo.get(issue.project.name, 0)
                if not sub_count:
                    minfo[issue.project.name] = 0
                minfo[issue.project.name] += 1
            self.moudle.append(minfo)

        # for project in self.moudle:
        #     bmodule = 0
        #     for key, value in project.items():
        #         if key in self.default_projects:
        #             bmodule = value
        #             print("\n%s, 问题单总数: %s\n" % (key, value))
        #         else:
        #             print("     - %s占比: %s%% \n" % (key, round(value / bmodule * 100, 2)))

    def individual(self):
        """统计任意人员任意时间段内的DI值数据"""
        for name, data in self.indivs.items():
            di = round(self.indivs[name]["DI"], 1)
            P15 = self.indivs[name]["Total"]
            P45, P45P = self.indivs[name]["P4-P5"], \
                        round(self.indivs[name]["P4-P5"] / self.indivs[name]["Total"] * 100, 2)
            invalid, invalid_pro = self.indivs[name]["Invalid"], \
                                   round(self.indivs[name]["Invalid"] / self.indivs[name]["Total"] * 100, 2)
            notestcase, notestcase_pro = self.indivs[name]["NoTestcase"], \
                                         round(self.indivs[name]["NoTestcase"] / self.indivs[name]["Total"] * 100, 2)
            nore, norepro = self.indivs[name]["NotReproduced"], \
                            round(self.indivs[name]["NotReproduced"] / self.indivs[name]["Total"] * 100, 2)
            res = [
                name, di, P15, P45, P45P, invalid, invalid_pro, notestcase, notestcase_pro, nore, norepro
            ]
            self.person.add_row(res)
            self.sheet2.append(res)


    @staticmethod
    def draw(xl, title):
        """调整列宽，表头调整背景色和字体"""
        # 背景色
        border_set = Border(
            left=Side(style='thin', color=colors.BLACK),
            right=Side(style='thin', color=colors.BLACK),
            top=Side(style='thin', color=colors.BLACK),
            bottom=Side(style='thin', color=colors.BLACK)
        )
        # 单元格格式
        for i, j in enumerate(title,1):
            xl.cell(row=1, column=i).fill = PatternFill("solid", fgColor="1F497D")
            xl.cell(row=1, column=i).font = Font(bold=True,color="FFFFFF")
            # xl.cell(row=1, column=j+1).alignment = Alignment(horizontal='center', vertical='center')
            xl.column_dimensions[get_column_letter(i)].width = str_len(j)
        for rows in xl:
            for cells in rows:
                cells.border = border_set




if __name__ == "__main__":
    print("\n---------------------------------------Redmine 问题单 DI 数值统计详情---------------------------------------")
    if os.path.exists("问题单库统计.xlsx"):
        os.remove("问题单库统计.xlsx")
    rm = RedmineDI()
    rm.di_statistics()
    print("-" * 150, "\n\n测试成员详细数据:\n")
    print(rm.person)
    print("\n", "-" * 150, "\n\n各产品线详细数据:\n")
    print(rm.project)
    rm.draw(rm.sheet1, rm.product_field_names), rm.draw(rm.sheet2, rm.team_field_names)
    rm.f.save("问题单库统计.xlsx")
